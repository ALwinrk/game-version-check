"""填充版本号脚本 — 根据已有的包名列爬取 version code 并写入.

用法:
    python scripts/fill_version_codes.py 海外游戏版本表.xlsx

    # 也可以指定代理
    set GVC_HTTP_PROXY=http://127.0.0.1:7897
    set GVC_HTTPS_PROXY=http://127.0.0.1:7897
    python scripts/fill_version_codes.py 海外游戏版本表.xlsx

    # 禁用慢速源（更快）
    set GVC_DISABLE_STEALTH=1
    python scripts/fill_version_codes.py 海外游戏版本表.xlsx

自动检测:
  - "当前后台版本号（vc)" 或 "当前后台版本号" 列 — 写入目标
  - "游戏包名" 或 "包名" 列 — 读取来源
  - 找不到目标列会自动在包名列右侧插入
"""

from __future__ import annotations

import sys
import os
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from gvc.config import MAX_GAME_WORKERS
from gvc.logging_setup import setup_logging, get_logger
from gvc.models import GameResult
from gvc.sources import query_all_sources
from gvc.version import best_version_code

setup_logging()
logger = get_logger()

# 列名匹配关键字（支持模糊匹配）
VC_COLUMN_KEYWORDS = ["当前后台版本号", "版本号(vc)", "version code", "vc"]
PKG_COLUMN_KEYWORDS = ["游戏包名", "包名", "package", "pkg", "packagename"]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _find_column(ws, keywords: list[str]) -> int | None:
    """在第 1 行表头中查找包含关键字的列，返回列号 (1-based)."""
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(1, col).value or "").lower().replace(" ", "")
        for kw in keywords:
            if kw.lower().replace(" ", "") in val:
                return col
    return None


def fill(filepath: str) -> None:
    """主逻辑."""
    logger.info("加载工作簿: %s", filepath)
    wb = load_workbook(filepath)
    ws = wb.active
    logger.info("工作表: %s (%d 行 × %d 列)", ws.title, ws.max_row, ws.max_column)

    # ── 1. 定位列 ──
    pkg_col = _find_column(ws, PKG_COLUMN_KEYWORDS)
    if not pkg_col:
        logger.error("未找到包名列！表头需包含: %s", " / ".join(PKG_COLUMN_KEYWORDS))
        logger.error("当前表头: %s", [str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)])
        return

    vc_col = _find_column(ws, VC_COLUMN_KEYWORDS)
    if not vc_col:
        # 自动在包名列右侧插入
        vc_col = pkg_col + 1
        logger.info("未找到版本号列，在 %s 列右侧插入", ws.cell(1, pkg_col).column_letter)
        ws.insert_cols(vc_col)
        cell = ws.cell(1, vc_col, "当前后台版本号（vc)")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    vc_letter = ws.cell(1, vc_col).column_letter
    pkg_letter = ws.cell(1, pkg_col).column_letter
    logger.info("包名列: %s (第 %d 列)", pkg_letter, pkg_col)
    logger.info("版本号列: %s (第 %d 列) — %s", vc_letter, vc_col, ws.cell(1, vc_col).value)

    # ── 2. 读取所有包名 ──
    rows_data: list[dict] = []
    for row in range(2, ws.max_row + 1):
        pkg = ws.cell(row, pkg_col).value
        if not pkg or not str(pkg).strip():
            continue
        rows_data.append({
            "row": row,
            "package": str(pkg).strip(),
        })

    logger.info("共 %d 款游戏待爬取", len(rows_data))
    if not rows_data:
        logger.error("未找到有效数据行")
        return

    # ── 3. 并发查询 ──
    results: dict[str, str] = {}
    with ThreadPoolExecutor(max_workers=min(len(rows_data), MAX_GAME_WORKERS)) as executor:
        future_map = {
            executor.submit(query_all_sources, d["package"]): d
            for d in rows_data
        }
        done = 0
        for future in as_completed(future_map):
            d = future_map[future]
            done += 1
            try:
                source_results = future.result()
            except Exception as e:
                logger.error("[%d/%d] %s 查询失败: %s", done, len(rows_data), d["package"], e)
                continue

            r = GameResult.from_source_results(
                package=d["package"],
                source_results=source_results,
            )
            vc = best_version_code(r)
            results[d["package"]] = vc
            logger.info("[%d/%d] %s → vc:%s", done, len(rows_data), d["package"], vc or "(无)")

    # ── 4. 写入 ──
    filled = 0
    for d in rows_data:
        vc = results.get(d["package"], "")
        if vc:
            cell = ws.cell(d["row"], vc_col, vc)
            cell.alignment = Alignment(horizontal="center")
            filled += 1
        elif ws.cell(d["row"], vc_col).value is None:
            # 没查到也标记一下，避免下次重复爬
            cell = ws.cell(d["row"], vc_col, "-")
            cell.alignment = Alignment(horizontal="center")

    # 设置列宽
    ws.column_dimensions[vc_letter].width = 18

    # ── 5. 保存 ──
    wb.save(filepath)
    logger.info("完成 — %d/%d 行已填充 version code，保存至 %s", filled, len(rows_data), filepath)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"用法: python {sys.argv[0]} <excel文件.xlsx>")
        print(f"示例: python {sys.argv[0]} C:\\Users\\Administrator\\Desktop\\海外游戏版本表.xlsx")
        sys.exit(1)
    fill(sys.argv[1])
