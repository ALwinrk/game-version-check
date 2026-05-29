"""一次性迁移脚本 v4→v5：添加版本号列并爬取数据.

将 Excel 从 v4 布局升级到 v5：
  1. 在 D 列（列 4）插入"当前后台版本号"，原有日期列自动右移
  2. 并发查询所有游戏的 version code（Google Play + 4 个 APK 站）
  3. 将最佳 version code 写入 D 列

用法:
    python scripts/migrate_to_v5.py 海外游戏版本表.xlsx

安全措施：
  - 检测 D 列表头是否已是"当前后台版本号"，避免重复插入
  - 只读不改原有列（除新增 D 列）
"""

from __future__ import annotations

import sys
import os
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# 确保 gvc 包可见
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from gvc.config import COL_CURRENT_VC, COL_NAME, COL_PACKAGE, MAX_GAME_WORKERS
from gvc.logging_setup import setup_logging, get_logger
from gvc.models import GameResult, SourceResult
from gvc.sources import query_all_sources
from gvc.version import best_version_code

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

NEW_COL = COL_CURRENT_VC  # 4 = D 列
HEADER_TEXT = "当前后台版本号"

setup_logging()
logger = get_logger()


def migrate(filepath: str) -> None:
    """执行迁移."""
    logger.info("加载工作簿: %s", filepath)
    wb = load_workbook(filepath)
    ws = wb.active

    # ── 1. 插入 D 列（安全检查） ──
    existing = ws.cell(1, NEW_COL).value
    if existing and HEADER_TEXT in str(existing):
        logger.warning("D 列已有版本号表头，跳过列插入（仅更新数据）")
    else:
        logger.info("在 D 列插入\"%s\"（原有日期列右移）...", HEADER_TEXT)
        ws.insert_cols(NEW_COL)

        # 写入表头
        cell = ws.cell(1, NEW_COL, HEADER_TEXT)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions[ws.cell(1, NEW_COL).column_letter].width = 18

    # ── 2. 读取所有包名 ──
    packages: list[dict] = []
    for row in range(2, ws.max_row + 1):
        pkg = ws.cell(row, COL_PACKAGE).value
        if not pkg or not str(pkg).strip():
            continue
        packages.append({
            "row": row,
            "package": str(pkg).strip(),
            "name": str(ws.cell(row, COL_NAME).value or "").strip(),
        })

    logger.info("共 %d 款游戏待爬取 version code", len(packages))
    if not packages:
        logger.error("未找到有效数据行")
        return

    # ── 3. 并发爬取 ──
    results: dict[str, str] = {}  # package → best version code
    with ThreadPoolExecutor(max_workers=min(len(packages), MAX_GAME_WORKERS)) as executor:
        future_map = {
            executor.submit(query_all_sources, d["package"]): d
            for d in packages
        }
        done = 0
        for future in as_completed(future_map):
            d = future_map[future]
            done += 1
            try:
                source_results = future.result()
            except Exception as e:
                logger.error("[%d/%d] %s 查询失败: %s", done, len(packages), d["package"], e)
                continue

            r = GameResult(
                package=d["package"],
                google=source_results.get("Google Play", SourceResult()),
                apkpure=source_results.get("APKPure", SourceResult()),
                apkcombo=source_results.get("APKCombo", SourceResult()),
                apkvision=source_results.get("APKVision", SourceResult()),
                apkmirror=source_results.get("APKMirror", SourceResult()),
                apkdl=source_results.get("APKDL", SourceResult()),
            )
            vc = best_version_code(r)
            results[d["package"]] = vc
            logger.info("[%d/%d] %s → vc:%s", done, len(packages), d["package"], vc or "(无)")

    # ── 4. 写入 D 列 ──
    filled = 0
    for d in packages:
        vc = results.get(d["package"], "")
        if vc:
            cell = ws.cell(d["row"], NEW_COL, vc)
            cell.alignment = Alignment(horizontal="center")
            filled += 1

    # ── 5. 保存 ──
    wb.save(filepath)
    logger.info("迁移完成 — %d/%d 行已填充 version code，保存至 %s", filled, len(packages), filepath)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"用法: python {sys.argv[0]} <excel文件.xlsx>")
        sys.exit(1)
    migrate(sys.argv[1])
