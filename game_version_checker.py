"""
游戏版本自动排查工具 v3
========================
导入 Excel 表格，自动排查 Google Play + 4 个 APK 站，将结果写回新列。
支持单独排查包名并记录版本变化历史。

用法:
  python game_version_checker.py 海外游戏表.xlsx          # 全表排查
  python game_version_checker.py --check com.tencent.ig    # 单独排查（含历史比对）
  python game_version_checker.py --check "pkg1,pkg2"       # 批量排查（含历史比对）
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

import requests
from bs4 import BeautifulSoup
from google_play_scraper import app as gp_app
try:
    from curl_cffi import requests as cf_requests
except ImportError:
    cf_requests = None  # curl_cffi 没装时降级
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ═══════════════════════════════════════════════════════════════════════
# 配置
# ═══════════════════════════════════════════════════════════════════════

REQUEST_DELAY   = 1.0
MAX_WORKERS     = 3
CF_TIMEOUT      = 12
REGULAR_TIMEOUT = 10

COL_NAME    = 1
COL_PACKAGE = 2
COL_CURRENT = 3

HEADER_FILL  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
UPDATE_FILL  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

VERSIONS_FILE = "last_versions.json"

# ═══════════════════════════════════════════════════════════════════════
# HTTP
# ═══════════════════════════════════════════════════════════════════════

_regular_session: requests.Session | None = None

def _get_session() -> requests.Session:
    global _regular_session
    if _regular_session is None:
        _regular_session = requests.Session()
        _regular_session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        })
    return _regular_session

def _http_get(url: str) -> tuple[int, str]:
    try:
        resp = _get_session().get(url, timeout=(5, REGULAR_TIMEOUT))
        if resp.status_code == 200 and len(resp.text) > 500:
            return resp.status_code, resp.text
        if resp.status_code != 403:
            return resp.status_code, resp.text
    except Exception:
        pass
    if cf_requests is not None:
        try:
            resp = cf_requests.get(url, impersonate="chrome124", timeout=CF_TIMEOUT)
            return resp.status_code, resp.text
        except Exception as e:
            return 0, str(e)[:200]
    return 0, "curl_cffi not installed"

# ═══════════════════════════════════════════════════════════════════════
# 数据结构
# ═══════════════════════════════════════════════════════════════════════

@dataclass
class SourceResult:
    version: str | None = None
    version_code: str | None = None
    updated_ts: int | None = None
    error: str | None = None

@dataclass
class GameResult:
    package: str
    name: str = ""
    google: SourceResult = field(default_factory=SourceResult)
    apkpure: SourceResult = field(default_factory=SourceResult)
    apkcombo: SourceResult = field(default_factory=SourceResult)
    apkvision: SourceResult = field(default_factory=SourceResult)
    apkmirror: SourceResult = field(default_factory=SourceResult)
    current_backend_version: str = ""
    has_update: bool = False
    update_detail: str = ""

# ═══════════════════════════════════════════════════════════════════════
# HTML 解析
# ═══════════════════════════════════════════════════════════════════════

def _extract_version(html: str) -> str | None:
    soup = BeautifulSoup(html, "html.parser")
    for elem in soup.select('[class*="version"]'):
        text = elem.get_text(strip=True)
        m = re.match(r'^[\d]+\.[\d]+(?:\.[\d]+)*$', text)
        if m and 3 < len(text) < 25: return text
    for attr in ("data-dt-version","data-version","data-app-version","data-release-version"):
        for elem in soup.find_all(attrs={attr: True}):
            v = elem[attr].strip()
            if re.match(r'^[\d]+\.[\d]+', v) and len(v) < 25: return v
    for elem in soup.select('[itemprop="version"]'):
        m = re.search(r'([\d]+\.[\d]+(?:\.[\d]+)?)', elem.get_text(strip=True))
        if m: return m.group(1)
    text = soup.get_text()
    m = re.search(r'(?:Version|v\.?)\s*:?\s*([\d]+\.[\d]+(?:\.[\d]+)?)', text, re.IGNORECASE)
    if m: return m.group(1)
    m = re.search(r'>\s*([\d]+\.[\d]+\.[\d]+)\s*<', html)
    if m: return m.group(1)
    return None

def _extract_version_code(html: str) -> str | None:
    m = re.search(r'variant\s*code[:\s]*(\d{6,12})', html, re.IGNORECASE)
    if m: return m.group(1)
    for m in re.finditer(r'(\d+\.\d+\.\d+)\s*[\(（]\s*(\d{6,12})\s*[\)）]', html):
        return m.group(2)
    m = re.search(r'data-versioncode\s*=\s*[\"\']?(\d{6,12})', html, re.IGNORECASE)
    if m: return m.group(1)
    return None

# ═══════════════════════════════════════════════════════════════════════
# 5 个数据源
# ═══════════════════════════════════════════════════════════════════════

def _check_google(package: str) -> SourceResult:
    try:
        info = gp_app(package, lang="en", country="us")
        version = info.get("version", "")
        updated = info.get("updated")
        if version.lower() in ("varies with device", "varies", ""):
            return SourceResult(error="Varies with device")
        return SourceResult(version=version, updated_ts=updated)
    except Exception as e:
        return SourceResult(error=str(e)[:100])

def _check_apk_site(url: str) -> SourceResult:
    try:
        status, html = _http_get(url)
        if status != 200:
            return SourceResult(error=f"HTTP {status}" if status else f"连接失败:{html[:40]}")
        version = _extract_version(html)
        vcode = _extract_version_code(html)
        if version:
            return SourceResult(version=version, version_code=vcode)
        return SourceResult(version_code=vcode, error=f"未匹配({len(html)}字节)")
    except Exception as e:
        return SourceResult(error=str(e)[:80])

def _check_apkpure(pkg: str)   -> SourceResult: return _check_apk_site(f"https://apkpure.com/search?q={pkg}")
def _check_apkcombo(pkg: str)  -> SourceResult: return _check_apk_site(f"https://apkcombo.com/search?q={pkg}")
def _check_apkvision(pkg: str) -> SourceResult: return _check_apk_site(f"https://apkvision.org/search?q={pkg}")
def _check_apkmirror(pkg: str) -> SourceResult: return _check_apk_site(f"https://www.apkmirror.com/?s={pkg}")

# ═══════════════════════════════════════════════════════════════════════
# 单游戏排查
# ═══════════════════════════════════════════════════════════════════════

def check_one(package: str) -> GameResult:
    r = GameResult(package=package)
    r.google    = _check_google(package);    time.sleep(REQUEST_DELAY)
    r.apkpure   = _check_apkpure(package);   time.sleep(REQUEST_DELAY)
    r.apkcombo  = _check_apkcombo(package);  time.sleep(REQUEST_DELAY)
    r.apkvision = _check_apkvision(package); time.sleep(REQUEST_DELAY)
    r.apkmirror = _check_apkmirror(package)
    return r

# ═══════════════════════════════════════════════════════════════════════
# 版本工具
# ═══════════════════════════════════════════════════════════════════════

def _normalize(v: str) -> str:
    if not v or v.lower() in ("varies with device", "varies"): return ""
    return re.sub(r'^[vV]\s*', '', re.sub(r'[\s\-_]+', '.', v.strip()))

def _best_version(r: GameResult) -> str:
    versions = [v for v in [r.google.version, r.apkpure.version, r.apkcombo.version,
                             r.apkvision.version, r.apkmirror.version] if v]
    if not versions: return "无法获取"
    counts = Counter(versions)
    top = counts.most_common(1)[0]
    if top[1] >= 2 or len(versions) == 1: return top[0]
    if r.google.version: return r.google.version
    return max(versions, key=lambda v: (len(v), v))

# ═══════════════════════════════════════════════════════════════════════
# Excel 结果文本
# ═══════════════════════════════════════════════════════════════════════

def _build_result_text(r: GameResult) -> str:
    best_v = _best_version(r)
    if best_v == "无法获取":
        r.update_detail = best_v; return "获取失败"

    best_vc = (r.google.version_code or r.apkpure.version_code or
               r.apkcombo.version_code or r.apkvision.version_code or
               r.apkmirror.version_code)
    current = _normalize(r.current_backend_version or "")

    if current and _normalize(best_v) != current:
        text = f"{r.current_backend_version}→{best_v}"
        if best_vc: text += f" (vc:{best_vc})"
        r.has_update = True; r.update_detail = text; return text

    r.update_detail = "-"; return "-"

# ═══════════════════════════════════════════════════════════════════════
# Excel 读写
# ═══════════════════════════════════════════════════════════════════════

def _find_or_create_date_column(ws, today_str: str) -> int:
    for col in range(COL_CURRENT + 1, ws.max_column + 1):
        if ws.cell(1, col).value == today_str:
            return col
    new_col = ws.max_column + 1
    cell = ws.cell(1, new_col, today_str)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    return new_col

def process_excel(filepath: str) -> None:
    print(f"读取表格: {filepath}")
    wb = load_workbook(filepath)
    ws = wb.active

    rows_data: list[dict] = []
    for row in range(2, ws.max_row + 1):
        pkg = ws.cell(row, COL_PACKAGE).value
        if not pkg or not str(pkg).strip(): continue
        rows_data.append({
            "row": row,
            "name": str(ws.cell(row, COL_NAME).value or "").strip(),
            "package": str(pkg).strip(),
            "current_version": str(ws.cell(row, COL_CURRENT).value or "").strip(),
        })

    if not rows_data:
        print("[错误] 表格中没有有效的包名数据"); return

    packages = [d["package"] for d in rows_data]
    print(f"共 {len(packages)} 个包名待排查\n")

    # 包名 → 当前后台版本/名称映射
    pkg_to_current = {d["package"]: d["current_version"] for d in rows_data}
    pkg_to_name = {d["package"]: d["name"] for d in rows_data}

    results: list[GameResult] = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_map = {executor.submit(check_one, p): p for p in packages}
        done = 0
        for future in as_completed(future_map):
            pkg = future_map[future]; done += 1
            try:
                r = future.result()
            except Exception as e:
                r = GameResult(package=pkg); r.google.error = str(e)
            r.current_backend_version = pkg_to_current.get(pkg, "")
            r.name = pkg_to_name.get(pkg, "")
            results.append(r)

            # 实时对比
            best = _best_version(r)
            cur = _normalize(r.current_backend_version)
            if not cur and best != "无法获取":
                flag = "[首次]"
            elif cur and _normalize(best) != cur and best != "无法获取":
                flag = f"⚠ {cur} → {best}"
            elif cur:
                flag = f"✓ (当前 {cur})"
            else:
                flag = ""
            print(f"  [{done}/{len(packages)}] {pkg} → {best}  {flag}")

    results.sort(key=lambda r: packages.index(r.package))

    for r in results:
        _build_result_text(r)

    today_str = datetime.now().strftime("%Y-%m-%d")
    date_col = _find_or_create_date_column(ws, today_str)

    updated_count = 0
    filled_count = 0
    for r, d in zip(results, rows_data):
        # 写入日期列
        cell = ws.cell(d["row"], date_col, r.update_detail)
        cell.alignment = Alignment(horizontal="center")
        if r.has_update:
            cell.fill = UPDATE_FILL; updated_count += 1
        if r.update_detail in ("-", "获取失败"):
            cell.fill = PatternFill()

        # 如果后台版本列为空，自动填充排查到的版本
        backend_cell = ws.cell(d["row"], COL_CURRENT)
        if not backend_cell.value or str(backend_cell.value).strip() == "":
            best = _best_version(r)
            if best != "无法获取":
                backend_cell.value = best
                filled_count += 1

    ws.column_dimensions[ws.cell(1, date_col).column_letter].width = 30
    wb.save(filepath)

    print(f"\n✅ 完成 — {updated_count}/{len(results)} 款有更新")
    if filled_count:
        print(f"   🆕 {filled_count} 款首次填充后台版本名")
    print(f"   结果写入: 列 {ws.cell(1, date_col).column_letter} ({today_str})")
    print(f"   文件已保存: {filepath}")
    if updated_count:
        print(f"\n⚠ 更新列表:")
        for r in results:
            if r.has_update: print(f"   {r.name or r.package}: {r.update_detail}")

# ═══════════════════════════════════════════════════════════════════════
# --check 模式 — 本地记录 + 版本比对
# ═══════════════════════════════════════════════════════════════════════

def _load_history() -> dict[str, dict]:
    if os.path.exists(VERSIONS_FILE):
        try:
            with open(VERSIONS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            result = {}
            for k, v in data.items():
                if isinstance(v, str):
                    result[k] = {"version": v}
                else:
                    result[k] = v
            return result
        except Exception:
            pass
    return {}

def _save_history(records: dict[str, dict]) -> None:
    with open(VERSIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

def check_packages(pkg_list: list[str], current_versions: list[str] | None = None) -> None:
    """单独排查并比对本地历史记录（可指定当前后台版本）"""
    if current_versions and len(current_versions) != len(pkg_list):
        print(f"[错误] --current 版本数量({len(current_versions)})与包名数量({len(pkg_list)})不一致")
        return

    print(f"排查 {len(pkg_list)} 个包名:\n")
    history = _load_history()
    new_history: dict[str, dict] = {}
    updated_list: list[tuple[str, str]] = []

    for i, pkg in enumerate(pkg_list, 1):
        print(f"[{i}/{len(pkg_list)}] {pkg} ...", end=" ", flush=True)
        r = check_one(pkg)
        best_v = _best_version(r)
        best_vc = (r.google.version_code or r.apkpure.version_code or
                   r.apkcombo.version_code or r.apkvision.version_code or
                   r.apkmirror.version_code)
        best_ts = r.google.updated_ts

        # 比对来源：优先用 --current，否则用历史记录
        if current_versions:
            compare_v = current_versions[i - 1].strip()
        else:
            last = history.get(pkg, {})
            compare_v = last.get("version", "") if isinstance(last, dict) else last

        changed = False; parts = []

        if compare_v and best_v != "无法获取" and _normalize(best_v) != _normalize(compare_v):
            changed = True; parts.append(f"{compare_v} → {best_v}")

        status = "⚠ 有更新" if changed else ("✓ 无变化" if compare_v else "  首次记录")
        print(f"{best_v}  {status}")

        for sname, s in [
            ("Google Play", r.google), ("APKPure", r.apkpure),
            ("APKCombo", r.apkcombo), ("APKVision", r.apkvision),
            ("APKMirror", r.apkmirror),
        ]:
            if s.version:
                detail = s.version
                if s.version_code: detail += f" (vc:{s.version_code})"
                if s.updated_ts and sname == "Google Play":
                    detail += f" [{datetime.fromtimestamp(s.updated_ts).strftime('%Y-%m-%d')}]"
                print(f"    {sname:<14} {detail}")
            elif s.error:
                print(f"    {sname:<14} ✗ {s.error[:50]}")

        if changed:
            detail = "; ".join(parts)
            print(f"  >>> {detail}")
            updated_list.append((pkg, detail))

        new_history[pkg] = {"version": best_v, "version_code": best_vc, "updated_ts": best_ts}
        print()

    _save_history(new_history)
    if updated_list:
        print(f"⚠ 发现 {len(updated_list)} 款有更新:")
        for pkg, detail in updated_list:
            print(f"   {pkg}: {detail}")
    else:
        print("✓ 所有包名版本无变化")

# ═══════════════════════════════════════════════════════════════════════
# 入口
# ═══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="游戏版本自动排查工具 v3")
    parser.add_argument("file", nargs="?", help="Excel 表格路径")
    parser.add_argument("--check", "-c", help="单独排查包名，多个用逗号分隔")
    parser.add_argument("--current", "-v", help="当前后台版本名，多个用逗号分隔（配合--check使用）")
    args = parser.parse_args()

    if args.check:
        pkgs = [p.strip() for p in args.check.split(",") if p.strip()]
        if not pkgs:
            print("请提供至少一个包名"); sys.exit(1)
        cur_vers = None
        if args.current:
            cur_vers = [v.strip() for v in args.current.split(",")]
        check_packages(pkgs, cur_vers)
    elif args.file:
        if not os.path.exists(args.file):
            print(f"[错误] 文件不存在: {args.file}"); sys.exit(1)
        process_excel(args.file)
    else:
        parser.print_help()
        print("\n示例:")
        print("  python game_version_checker.py 海外游戏表.xlsx")
        print("  python game_version_checker.py --check com.tencent.ig")
        print("  python game_version_checker.py --check com.a,com.b --current 4.0,6.5")
        print("  python game_version_checker.py -c \"com.a,com.b\"")

if __name__ == "__main__":
    main()
