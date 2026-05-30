"""Excel 读写处理."""

from __future__ import annotations

from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from gvc.config import COL_CURRENT, COL_CURRENT_VC, COL_NAME, COL_PACKAGE
from gvc.logging_setup import get_logger
from gvc.models import GameResult, SourceResult
from gvc.version import best_version, normalize

logger = get_logger()

# ── 样式 ─────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
UPDATE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def read_rows(ws: Worksheet) -> list[dict]:
    """从工作表读取所有有效数据行.

    Returns:
        [{row, name, package, current_version}, ...]
    """
    rows: list[dict] = []
    for row in range(2, ws.max_row + 1):
        pkg = ws.cell(row, COL_PACKAGE).value
        if not pkg or not str(pkg).strip():
            continue
        rows.append({
            "row": row,
            "name": str(ws.cell(row, COL_NAME).value or "").strip(),
            "package": str(pkg).strip(),
            "current_version": str(ws.cell(row, COL_CURRENT).value or "").strip(),
            "current_version_code": str(ws.cell(row, COL_CURRENT_VC).value or "").strip(),
        })
    return rows


def find_or_create_date_column(ws: Worksheet, today_str: str) -> int:
    """查找或创建今日日期列."""
    for col in range(COL_CURRENT_VC + 1, ws.max_column + 1):
        if ws.cell(1, col).value == today_str:
            return col

    new_col = ws.max_column + 1
    cell = ws.cell(1, new_col, today_str)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    return new_col


def build_result_text(r: GameResult) -> str:
    """根据 GameResult 生成 Excel 单元格文本.

    对比策略（v5）：
    1. 优先版本号对比 — cur_vc < best_vc → 有更新
    2. 无版本号时回退版本名对比
    """
    from gvc.version import best_version_code as _best_vc, compare_version_codes

    best_v = best_version(r)
    if best_v == "无法获取":
        r.update_detail = "获取失败"
        r.has_update = False
        return "获取失败"

    best_vc = _best_vc(r)
    cur_vc = (r.current_backend_version_code or "").strip()
    cur_vn = normalize(r.current_backend_version or "")
    best_vn = normalize(best_v)

    # ── 策略 1：版本号对比（优先） ──
    if cur_vc and best_vc:
        cmp = compare_version_codes(cur_vc, best_vc)
        if cmp < 0:
            # 版本号有更新
            text = f"vc:{cur_vc}→{best_vc}"
            if r.current_backend_version and best_vn != cur_vn:
                text += f" ({r.current_backend_version}→{best_v})"
            r.has_update = True
            r.update_detail = text
            return text
        elif cmp == 0:
            # 版本号相同，检查版本名是否不同（极端情况）
            if cur_vn and best_vn != cur_vn:
                text = f"{r.current_backend_version}→{best_v}"
                if best_vc:
                    text += f" (vc:{best_vc})"
                r.has_update = True
                r.update_detail = text
                return text
            r.update_detail = "-"
            r.has_update = False
            return "-"

    # ── 策略 2：回退版本名对比 ──
    if cur_vn and best_vn != cur_vn:
        text = f"{r.current_backend_version}→{best_v}"
        if best_vc:
            text += f" (vc:{best_vc})"
        r.has_update = True
        r.update_detail = text
        return text

    r.update_detail = "-"
    r.has_update = False
    return "-"


def write_results(
    ws: Worksheet,
    rows_data: list[dict],
    results: list[GameResult],
    today_str: str,
) -> tuple[int, int]:
    """将结果写回工作表.

    Returns:
        (updated_count, filled_count)
    """
    date_col = find_or_create_date_column(ws, today_str)
    updated_count = 0
    filled_count = 0

    for r, d in zip(results, rows_data):
        # 写入日期列
        cell = ws.cell(d["row"], date_col, r.update_detail)
        cell.alignment = Alignment(horizontal="center")
        if r.has_update:
            cell.fill = UPDATE_FILL
            updated_count += 1
        elif r.update_detail in ("-", "获取失败"):
            cell.fill = PatternFill()

        # 自动填充空的后台版本列
        backend_cell = ws.cell(d["row"], COL_CURRENT)
        if not backend_cell.value or str(backend_cell.value).strip() == "":
            best_v = best_version(r)
            if best_v != "无法获取":
                backend_cell.value = best_v
                filled_count += 1

        # 自动填充空的版本号列
        from gvc.version import best_version_code as _best_vc
        vc_cell = ws.cell(d["row"], COL_CURRENT_VC)
        if not vc_cell.value or str(vc_cell.value).strip() == "":
            best_vc = _best_vc(r)
            if best_vc:
                vc_cell.value = best_vc

    # 调整列宽
    col_letter = ws.cell(1, date_col).column_letter
    ws.column_dimensions[col_letter].width = 30
    # 版本号列宽
    vc_col_letter = ws.cell(1, COL_CURRENT_VC).column_letter
    if ws.column_dimensions[vc_col_letter].width is None or ws.column_dimensions[vc_col_letter].width < 15:
        ws.column_dimensions[vc_col_letter].width = 18

    return updated_count, filled_count


def process_excel(filepath: str) -> None:
    """完整的 Excel 排查流程."""
    logger.info("读取表格: %s", filepath)

    wb = load_workbook(filepath)
    ws = wb.active

    rows_data = read_rows(ws)
    if not rows_data:
        logger.error("表格中没有有效的包名数据")
        return

    packages = [d["package"] for d in rows_data]
    logger.info("共 %d 个包名待排查", len(packages))

    # 延时导入避免循环依赖
    from concurrent.futures import ThreadPoolExecutor, as_completed

    from gvc.config import MAX_GAME_WORKERS
    from gvc.sources import query_all_sources

    pkg_to_current = {d["package"]: d["current_version"] for d in rows_data}
    pkg_to_current_vc = {d["package"]: d.get("current_version_code", "") for d in rows_data}
    pkg_to_name = {d["package"]: d["name"] for d in rows_data}

    results: list[GameResult] = []
    with ThreadPoolExecutor(max_workers=min(len(packages), MAX_GAME_WORKERS)) as executor:
        future_map = {
            executor.submit(query_all_sources, p): p
            for p in packages
        }
        done = 0
        for future in as_completed(future_map):
            pkg = future_map[future]
            done += 1
            try:
                source_results = future.result()
            except Exception as e:
                logger.error("查询 %s 失败: %s", pkg, e)
                source_results = {}

            r = GameResult(
                package=pkg,
                name=pkg_to_name.get(pkg, ""),
                current_backend_version=pkg_to_current.get(pkg, ""),
                current_backend_version_code=pkg_to_current_vc.get(pkg, ""),
                google=source_results.get("Google Play", SourceResult()),
                apkpure=source_results.get("APKPure", SourceResult()),
                apkcombo=source_results.get("APKCombo", SourceResult()),
                apkvision=source_results.get("APKVision", SourceResult()),
                apkmirror=source_results.get("APKMirror", SourceResult()),
                apkdl=source_results.get("APKDL", SourceResult()),
            )
            results.append(r)

            # 实时对比
            from gvc.version import best_version_code as _best_vc, compare_version_codes
            best_v = best_version(r)
            best_vc = _best_vc(r)
            cur_vn = normalize(r.current_backend_version)
            cur_vc = (r.current_backend_version_code or "").strip()

            if not cur_vn and best_v != "无法获取":
                flag = "[首次]"
            elif best_vc and cur_vc:
                cmp = compare_version_codes(cur_vc, best_vc)
                if cmp < 0:
                    flag = f"⚠ vc:{cur_vc}→{best_vc}"
                elif cmp == 0:
                    flag = f"✓ (当前 {cur_vn}, vc:{cur_vc})"
                else:
                    flag = ""
            elif cur_vn and normalize(best_v) != cur_vn and best_v != "无法获取":
                flag = f"⚠ {cur_vn}→{best_v}"
            elif cur_vn:
                flag = f"✓ (当前 {cur_vn})"
            else:
                flag = ""
            logger.info("[%d/%d] %s → %s  %s", done, len(packages), pkg, best_v, flag)

    # 保持原始顺序
    results.sort(key=lambda r: packages.index(r.package))

    # 生成结果文本
    for r in results:
        build_result_text(r)

    today_str = datetime.now().strftime("%Y-%m-%d")
    updated_count, filled_count = write_results(ws, rows_data, results, today_str)
    wb.save(filepath)

    logger.info("✅ 完成 — %d/%d 款有更新", updated_count, len(results))
    if filled_count:
        logger.info("   🆕 %d 款首次填充后台版本名", filled_count)
    logger.info("   结果写入: 列 %s (%s)", ws.cell(1, find_or_create_date_column(ws, today_str)).column_letter, today_str)
    logger.info("   文件已保存: %s", filepath)

    if updated_count:
        logger.info("⚠ 更新列表:")
        for r in results:
            if r.has_update:
                logger.info("   %s: %s", r.name or r.package, r.update_detail)
