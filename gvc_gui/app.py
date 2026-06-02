"""主应用窗口 — 现代化 GUI (v5 redesign)."""

from __future__ import annotations

import os
import queue
import tkinter as tk
from datetime import datetime
from tkinter import messagebox

import customtkinter as ctk

from gvc.logging_setup import setup_logging, get_logger
from gvc.excel_handler import read_rows, write_results
from gvc.models import GameResult

from gvc_gui.config import (
    WINDOW_TITLE, DEFAULT_GEOMETRY, MIN_WIDTH, MIN_HEIGHT,
    APPEARANCE_MODE, COLOR_THEME, POLL_INTERVAL,
    FONT_NORMAL, FONT_SMALL, FONT_BOLD, FONT_CAPTION,
    BG_PRIMARY, TEXT_PRIMARY, TEXT_SECONDARY, TEXT_DISABLED,
    STATUS_BAR_BG, SURFACE_BG, ACCENT_SUCCESS,
)
from gvc_gui.main_view import _resolve
from gvc_gui.dialogs import (
    SettingsDialog, SingleCheckDialog, AboutDialog,
    load_settings, save_settings,
)
from gvc_gui.main_view import StatsDashboard, FilePickerFrame, GameTableFrame
from gvc_gui.progress_view import ProgressView
from gvc_gui.workers import CheckWorker, WorkerMessage

setup_logging()
logger = get_logger()


class MainApplication(ctk.CTk):
    """主应用程序窗口."""

    def __init__(self) -> None:
        super().__init__()

        # ── 设置 ──
        self._settings = load_settings()
        ctk.set_appearance_mode(self._settings.get("appearance_mode", APPEARANCE_MODE))
        ctk.set_default_color_theme(self._settings.get("color_theme", COLOR_THEME))

        # ── 启动时应用代理/超时等运行时设置 ──
        self._apply_runtime_settings()

        # ── 窗口 ──
        self.title(WINDOW_TITLE)
        self.geometry(DEFAULT_GEOMETRY)
        self.minsize(MIN_WIDTH, MIN_HEIGHT)
        self.configure(fg_color=BG_PRIMARY)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        # ── 状态 ──
        self._worker: CheckWorker | None = None
        self._msg_queue: queue.Queue = queue.Queue()
        self._results: dict[str, GameResult] = {}
        self._rows_data: list[dict] = []
        self._current_file: str = ""
        self._running: bool = False

        # ── 构建 UI ──
        self._build_menu()
        self._build_layout()
        self._poll_queue()

        # 网格权重
        self.grid_rowconfigure(3, weight=1)  # 表格区可伸展
        self.grid_columnconfigure(0, weight=1)

    # ══════════════════════════════════════════════════════════
    # 菜单
    # ══════════════════════════════════════════════════════════

    def _build_menu(self) -> None:
        is_dark = ctk.get_appearance_mode() == "Dark"
        menu_bg = "#1E293B" if is_dark else "#FFFFFF"
        menu_fg = "#F1F5F9" if is_dark else "#0F172A"
        menu_active_bg = "#334155" if is_dark else "#E2E8F0"
        menu_active_fg = "#FFFFFF" if is_dark else "#0F172A"

        menubar = tk.Menu(
            self, font=FONT_SMALL,
            bg=menu_bg, fg=menu_fg,
            activebackground=menu_active_bg,
            activeforeground=menu_active_fg,
            borderwidth=0, relief="flat",
            activeborderwidth=0,
        )

        def _submenu() -> tk.Menu:
            return tk.Menu(
                menubar, tearoff=0,
                bg=menu_bg, fg=menu_fg,
                activebackground=menu_active_bg,
                activeforeground=menu_active_fg,
                font=FONT_SMALL,
            )

        file_menu = _submenu()
        file_menu.add_command(label=" 打开表格…", command=self._on_open_file)
        file_menu.add_command(label=" 保存结果到 Excel", command=self._on_save_results)
        file_menu.add_separator()
        file_menu.add_command(label=" 退出", command=self._on_close)
        menubar.add_cascade(label=" 文件 ", menu=file_menu)

        check_menu = _submenu()
        check_menu.add_command(label=" 单独排查…", command=self._on_single_check)
        menubar.add_cascade(label=" 排查 ", menu=check_menu)

        settings_menu = _submenu()
        settings_menu.add_command(label=" 选项…", command=self._on_settings)
        menubar.add_cascade(label=" 设置 ", menu=settings_menu)

        help_menu = _submenu()
        help_menu.add_command(label=" 关于", command=self._on_about)
        menubar.add_cascade(label=" 帮助 ", menu=help_menu)

        self.configure(menu=menubar)

    # ══════════════════════════════════════════════════════════
    # 布局
    # ══════════════════════════════════════════════════════════

    def _build_layout(self) -> None:
        # 主窗口用 grid: row 0=滚动区, row 1=状态栏
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # ═══ 可滚动内容区（含全部组件） ═══
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.grid(row=0, column=0, sticky="nsew", padx=4, pady=(4, 0))
        scroll.grid_columnconfigure(0, weight=1)

        # ROW 0 — 统计仪表盘
        self._stats = StatsDashboard(scroll)
        self._stats.grid(row=0, column=0, sticky="ew", padx=4, pady=(2, 0))

        # ROW 1 — 双面板：表格排查 + 单独排查
        row1 = ctk.CTkFrame(scroll, fg_color="transparent")
        row1.grid(row=1, column=0, sticky="ew", padx=4, pady=(6, 0))
        row1.grid_columnconfigure(0, weight=3)
        row1.grid_columnconfigure(1, weight=2)

        # 左侧：表格排查
        left_panel = ctk.CTkFrame(row1, fg_color="transparent")
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 4))
        left_panel.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(left_panel, text="📂 批量表格排查",
                     font=FONT_BOLD, text_color=TEXT_PRIMARY).pack(anchor="w")
        self._file_picker = FilePickerFrame(left_panel)
        self._file_picker.pack(fill="x", pady=(4, 0))
        self._file_picker.set_callbacks(
            on_start=self._start_check,
            on_stop=self._stop_check,
        )

        # 右侧：单独排查
        right_panel = ctk.CTkFrame(row1, fg_color=SURFACE_BG, corner_radius=8)
        right_panel.grid(row=0, column=1, sticky="nsew", padx=(4, 0))
        self._build_single_check_panel(right_panel)

        # ROW 2 — 进度 + 日志
        self._progress_view = ProgressView(scroll, height=130)
        self._progress_view.grid(row=2, column=0, sticky="ew", padx=4, pady=(4, 0))

        # ROW 3 — 游戏表格
        self._table = GameTableFrame(scroll)
        self._table.grid(row=3, column=0, sticky="nsew", padx=4, pady=(4, 4))
        self._table.set_callbacks(
            on_recheck=self._on_recheck_selected,
            on_copy=lambda: None,
        )

        # ═══ 状态栏（固定底部，不滚动） ═══
        status_frame = ctk.CTkFrame(
            self, height=32, corner_radius=0,
            fg_color=STATUS_BAR_BG,
        )
        status_frame.grid(row=1, column=0, sticky="ew")
        status_frame.grid_propagate(False)

        # 左侧：状态圆点 + 文字
        left = ctk.CTkFrame(status_frame, fg_color="transparent")
        left.pack(side="left", fill="y", padx=(12, 0))

        self._status_dot = ctk.CTkLabel(
            left, text="●", font=("", 10),
            text_color=("#64748B", "#94A3B8"),
            width=16,
        )
        self._status_dot.pack(side="left")

        # 检测代理配置
        proxy_configured = bool(self._settings.get("http_proxy") or self._settings.get("https_proxy"))
        ready_text = " 就绪 | 选择表格或输入包名开始排查"
        if not proxy_configured:
            ready_text += " | ⚠ 未配置代理 (海外站点可能无法连接，请打开 设置→选项)"
        self._status_label = ctk.CTkLabel(
            left, text=ready_text,
            font=FONT_SMALL, anchor="w",
            text_color=TEXT_SECONDARY,
        )
        self._status_label.pack(side="left", fill="x")

        # 右侧：外观模式
        right = ctk.CTkFrame(status_frame, fg_color="transparent")
        right.pack(side="right", padx=(0, 12))
        self._mode_label = ctk.CTkLabel(
            right, text="", font=FONT_CAPTION, text_color=TEXT_DISABLED,
        )
        self._mode_label.pack(side="right")

        ctk.AppearanceModeTracker.add(self._on_appearance_changed, self)
        self._update_mode_indicator()

    def _build_single_check_panel(self, parent: ctk.CTkFrame) -> None:
        """构建右侧「单独排查」面板."""
        # 标题
        ctk.CTkLabel(parent, text="🔍 单独排查",
                     font=FONT_BOLD, text_color=TEXT_PRIMARY).pack(anchor="w", padx=12, pady=(10, 4))

        # 包名
        ctk.CTkLabel(parent, text="游戏包名",
                     font=FONT_SMALL, text_color=TEXT_SECONDARY).pack(anchor="w", padx=12)
        self._sc_pkg_var = ctk.StringVar()
        ctk.CTkEntry(parent, textvariable=self._sc_pkg_var,
                     font=FONT_NORMAL, height=32,
                     placeholder_text="例: com.tencent.ig").pack(fill="x", padx=12, pady=(0, 4))

        # 当前版本名（可选）
        ctk.CTkLabel(parent, text="后台版本名（可选）",
                     font=FONT_SMALL, text_color=TEXT_SECONDARY).pack(anchor="w", padx=12)
        self._sc_ver_var = ctk.StringVar()
        ctk.CTkEntry(parent, textvariable=self._sc_ver_var,
                     font=FONT_NORMAL, height=32,
                     placeholder_text="例: 4.4.0").pack(fill="x", padx=12, pady=(0, 4))

        # 当前版本号（可选）
        ctk.CTkLabel(parent, text="后台版本号（可选）",
                     font=FONT_SMALL, text_color=TEXT_SECONDARY).pack(anchor="w", padx=12)
        self._sc_vc_var = ctk.StringVar()
        ctk.CTkEntry(parent, textvariable=self._sc_vc_var,
                     font=FONT_NORMAL, height=32,
                     placeholder_text="例: 12345678").pack(fill="x", padx=12, pady=(0, 4))

        # 下载勾选框
        dl_row = ctk.CTkFrame(parent, fg_color="transparent")
        dl_row.pack(fill="x", padx=12, pady=(2, 2))
        self._sc_download_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(dl_row, text="下载 64 位 APK", variable=self._sc_download_var,
                        font=FONT_SMALL).pack(side="left")

        # 按钮行
        btn_row = ctk.CTkFrame(parent, fg_color="transparent")
        btn_row.pack(fill="x", padx=12, pady=(4, 0))
        self._sc_check_btn = ctk.CTkButton(
            btn_row, text="排查", command=self._on_single_check_inline,
            font=FONT_NORMAL, height=32,
        )
        self._sc_check_btn.pack(side="left", fill="x", expand=True)
        self._sc_dl_btn = ctk.CTkButton(
            btn_row, text="下载", command=self._on_single_download,
            font=FONT_NORMAL, height=32,
            fg_color=ACCENT_SUCCESS, hover_color=("#16A34A", "#22C55E"),
            state="disabled",
        )
        self._sc_dl_btn.pack(side="left", padx=(4, 0))

        # 结果展示
        self._sc_result = ctk.CTkTextbox(parent, font=FONT_SMALL, height=100)
        self._sc_result.pack(fill="both", expand=True, padx=12, pady=(6, 10))
        self._sc_result.tag_config("heading", foreground=_resolve(("#3B82F6", "#60A5FA")))
        self._sc_result.tag_config("ok", foreground=_resolve(("#22C55E", "#4ADE80")))
        self._sc_result.tag_config("err", foreground=_resolve(("#EF4444", "#F87171")))
        self._sc_result.tag_config("warn", foreground=_resolve(("#F59E0B", "#FBBF24")))
        self._sc_result.insert("1.0", "输入包名点「排查」开始…")
        self._sc_result.configure(state="disabled")

        # 状态：保存上次查询结果用于下载
        self._sc_last_results: dict = {}
        self._sc_last_best_version: str = ""
        self._sc_last_package: str = ""

    # ══════════════════════════════════════════════════════════
    # 消息轮询
    # ══════════════════════════════════════════════════════════

    def _poll_queue(self) -> None:
        try:
            while True:
                msg: WorkerMessage = self._msg_queue.get_nowait()
                self._handle_message(msg)
        except queue.Empty:
            pass
        self.after(POLL_INTERVAL, self._poll_queue)

    def _handle_message(self, msg: WorkerMessage) -> None:
        if msg.type == "result" and msg.game_result:
            r = msg.game_result
            self._results[r.package] = r
            self._table.update_row(msg.index, r)  # worker 现在发原始行索引(0-based)
            self._progress_view.log_result(
                r.package, r.name,
                r.update_detail or "",
                r.update_detail or "",
                r.has_update,
            )
            self._progress_view.update_progress(msg.index + 1, msg.total, r.package)
            self._status_label.configure(
                text=f" [{msg.index + 1}/{msg.total}] {r.name or r.package}"
            )
            # 实时更新统计
            s = self._table.get_stats()
            self._stats.update(
                total=s["total"],
                updated=s["update"],
                ok=s["ok"],
                error=s["error"],
            )
        elif msg.type == "error":
            self._progress_view.log_error(msg.package, msg.error_text)
            self._set_status_dot("error")
            self._status_label.configure(
                text=f" [{msg.index + 1}/{msg.total}] {msg.package} — 查询失败"
            )
        elif msg.type == "done":
            self._on_done()

    # ══════════════════════════════════════════════════════════
    # 排查流程
    # ══════════════════════════════════════════════════════════

    def _start_check(self) -> None:
        filepath = self._file_picker.get_path()
        if not filepath:
            messagebox.showwarning("提示", "请先选择 Excel 表格")
            return
        if not os.path.exists(filepath):
            messagebox.showerror("错误", f"文件不存在:\n{filepath}")
            return

        self._current_file = filepath

        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active
            rows_data = read_rows(ws)
            wb.close()
        except Exception as e:
            messagebox.showerror("错误", f"无法读取 Excel:\n{e}")
            return

        if not rows_data:
            messagebox.showwarning("提示", "表格中没有有效的包名数据")
            return

        self._rows_data = rows_data
        self._results.clear()

        self._running = True
        self._file_picker.set_running(True)
        self._file_picker.set_info(f"共 {len(rows_data)} 款游戏")
        self._table.populate(rows_data)
        self._progress_view.reset(len(rows_data))
        self._stats.reset()
        self._set_status_dot("active")
        self._status_label.configure(text=f" 开始排查… 共 {len(rows_data)} 款游戏")

        max_workers = self._settings.get("max_game_workers", 3)
        self._worker = CheckWorker(rows_data, self._msg_queue, max_game_workers=max_workers)
        self._worker.start()

    def _stop_check(self) -> None:
        if self._worker:
            self._worker.cancel()
        self._running = False
        self._file_picker.set_running(False)
        self._set_status_dot("muted")
        self._status_label.configure(text=" 已手动停止")

    def _on_done(self) -> None:
        self._running = False
        self._file_picker.set_running(False)

        s = self._table.get_stats()
        updated = s["update"]
        failed = s["error"]
        total = s["total"]
        self._set_status_dot("success" if failed == 0 else "muted")
        self._progress_view.set_done(total, updated, failed)

        msg = f" 完成 — 共 {total} 款"
        if updated:
            msg += f" | {updated} 款有更新"
        if failed:
            msg += f" | {failed} 款失败"
        msg += " | 文件 → 保存结果到 Excel"
        self._status_label.configure(text=msg)

        # 自动弹出保存提示
        if updated > 0:
            if messagebox.askyesno("排查完成", f"共排查 {total} 款游戏\n{updated} 款有更新\n\n是否保存结果到 Excel？"):
                self._on_save_results()

    # ══════════════════════════════════════════════════════════
    # 文件 / 保存
    # ══════════════════════════════════════════════════════════

    def _on_open_file(self) -> None:
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="选择 Excel 表格",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if path:
            self._file_picker.set_path(path)
            try:
                from openpyxl import load_workbook
                wb = load_workbook(path)
                ws = wb.active
                rows = read_rows(ws)
                wb.close()
                self._file_picker.set_info(f"共 {len(rows)} 款游戏 | 准备好开始排查")
                self._stats.reset()
                self._status_label.configure(
                    text=f" 已加载: {os.path.basename(path)} | {len(rows)} 款游戏待排查"
                )
            except Exception:
                self._status_label.configure(text=f" 已选择: {os.path.basename(path)}")

    def _on_save_results(self) -> None:
        if not self._results:
            messagebox.showwarning("提示", "没有可保存的结果，请先执行排查")
            return
        if not self._current_file or not os.path.exists(self._current_file):
            messagebox.showerror("错误", "原文件不存在")
            return

        try:
            from openpyxl import load_workbook
            from gvc.excel_handler import build_result_text

            wb = load_workbook(self._current_file)
            ws = wb.active

            for d in self._rows_data:
                pkg = d["package"]
                r = self._results.get(pkg)
                if r:
                    build_result_text(r)

            today_str = datetime.now().strftime("%Y-%m-%d")
            write_results(
                ws, self._rows_data,
                [self._results.get(d["package"], GameResult(package=d["package"]))
                 for d in self._rows_data],
                today_str,
            )
            wb.save(self._current_file)
            wb.close()
            self._status_label.configure(text=f" 结果已保存 → {today_str} 列")
            messagebox.showinfo("保存完成", f"排查结果已写入 Excel「{today_str}」列")
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    # ══════════════════════════════════════════════════════════
    # 对话框
    # ══════════════════════════════════════════════════════════

    def _on_single_check(self) -> None:
        SingleCheckDialog(self)

    def _on_single_check_inline(self) -> None:
        """主界面右侧面板「单独排查」按钮."""
        pkg = self._sc_pkg_var.get().strip()
        if not pkg:
            messagebox.showwarning("提示", "请输入游戏包名")
            return

        self._sc_check_btn.configure(text="查询中…", state="disabled")
        self._sc_dl_btn.configure(state="disabled")
        self._sc_result.configure(state="normal")
        self._sc_result.delete("1.0", "end")
        self._sc_result.insert("1.0", "正在查询…")
        self._sc_result.configure(state="disabled")
        self._set_status_dot("active")
        self._status_label.configure(text=f" 单独排查: {pkg} …")

        import threading

        def _run() -> None:
            from gvc.models import GameResult, SourceResult
            from gvc.sources import query_all_sources
            from gvc.version import best_version, best_version_code, check_for_update

            # 勾选下载时强制查全部 6 个源（需要 detail_url 做下载链接）
            force_all = self._sc_download_var.get()
            results = query_all_sources(pkg, force_all=force_all)
            r = GameResult.from_source_results(pkg, results)
            bv = best_version(r)
            bvc = best_version_code(r)

            # 保存供后续下载
            self._sc_last_results = results
            self._sc_last_best_version = bv
            self._sc_last_package = pkg

            # 构建结果文本
            lines = [f"包名: {pkg}"]
            if bv != "无法获取":
                lines.append(f"最佳版本: {bv}" + (f" (vc:{bvc})" if bvc else ""))
            else:
                lines.append("最佳版本: 无法获取")
            lines.append("")

            cur_v = self._sc_ver_var.get().strip()
            cur_vc = self._sc_vc_var.get().strip()
            has_update = False
            if cur_v or cur_vc:
                has_update, detail = check_for_update(bv, bvc, cur_v, cur_vc)
                if has_update:
                    lines.append(f">>> {detail}")
                else:
                    lines.append("状态: ✓ 无变化")
                lines.append("")

            for name, s in results.items():
                if s.version:
                    line = f"  {name}: {s.version}"
                    if s.version_code:
                        line += f" (vc:{s.version_code})"
                elif s.error:
                    line = f"  {name}: ✗ {s.error[:40]}"
                else:
                    line = f"  {name}: (无数据)"
                lines.append(line)

            self.after(0, lambda: self._show_inline_result("\n".join(lines), has_update))

        threading.Thread(target=_run, daemon=True).start()

    def _show_inline_result(self, text: str, has_update: bool) -> None:
        """显示内嵌面板的排查结果."""
        self._sc_result.configure(state="normal")
        self._sc_result.delete("1.0", "end")
        for i, line in enumerate(text.split("\n")):
            if i < 2:
                self._sc_result.insert("end", line + "\n", "heading")
            elif line.startswith(">>>"):
                self._sc_result.insert("end", line + "\n", "warn")
            elif "✗" in line:
                self._sc_result.insert("end", line + "\n", "err")
            elif line.startswith("  ") and ":" in line:
                self._sc_result.insert("end", line + "\n", "ok")
            else:
                self._sc_result.insert("end", line + "\n")
        self._sc_result.configure(state="disabled")
        self._sc_check_btn.configure(text="排查", state="normal")

        # 启用下载按钮
        if has_update and self._sc_download_var.get():
            self._sc_dl_btn.configure(state="normal")
        elif self._sc_download_var.get() and self._sc_last_best_version not in ("", "无法获取"):
            self._sc_dl_btn.configure(state="normal")
        self._set_status_dot("success")
        self._status_label.configure(text=f" 单独排查完成: {self._sc_last_package}")

    def _on_single_download(self) -> None:
        """主界面右侧面板「下载」按钮 — 调起下载管理器."""
        if not self._sc_last_results or not self._sc_last_package:
            messagebox.showwarning("提示", "请先执行排查")
            return

        dm_name = self._settings.get("download_manager", "auto")
        dl_dir = self._settings.get("download_dir", "") or None

        self._sc_dl_btn.configure(text="下载中…", state="disabled")
        self._status_label.configure(text=f" 提取下载链接: {self._sc_last_package} …")

        import threading

        def _run() -> None:
            from gvc.downloader import auto_download
            result = auto_download(
                package=self._sc_last_package,
                source_results=self._sc_last_results,
                best_version=self._sc_last_best_version,
                download_dir=dl_dir,
                dm_name=dm_name,
            )
            self.after(0, lambda: self._show_download_result(result))

        threading.Thread(target=_run, daemon=True).start()

    def _show_download_result(self, result: dict) -> None:
        """显示下载结果."""
        self._sc_dl_btn.configure(text="下载", state="normal")
        if result.get("success"):
            mgr = result.get("manager", "")
            if mgr:
                msg = f"已调起 {mgr}\n{result.get('arch', '')} | {result.get('source', '')}"
                self._status_label.configure(text=f" 下载已发送: {self._sc_last_package}")
            else:
                msg = f"内置下载完成\n{result.get('file_path', '')}"
                self._status_label.configure(text=f" 下载完成: {self._sc_last_package}")
            messagebox.showinfo("下载", msg)
        elif result.get("only_32bit_urls"):
            urls = "\n".join(result["only_32bit_urls"][:3])
            messagebox.showinfo("仅 32 位 APK", f"不自动下载，链接如下:\n{urls}")
        else:
            messagebox.showwarning("下载失败", result.get("error", "未知错误"))
        self._sc_dl_btn.configure(state="normal" if self._sc_last_results else "disabled")

    def _on_settings(self) -> None:
        SettingsDialog(self, self._settings, self._apply_settings)

    def _apply_runtime_settings(self) -> None:
        """启动时应用已保存的代理/超时/并发/下载等运行时设置到环境变量."""
        for key, env in [("http_proxy", "GVC_HTTP_PROXY"), ("https_proxy", "GVC_HTTPS_PROXY")]:
            val = self._settings.get(key, "")
            std_env = "HTTP_PROXY" if key == "http_proxy" else "HTTPS_PROXY"
            if val:
                os.environ[env] = val
                os.environ[std_env] = val
            else:
                for v in (env, std_env):
                    if v in os.environ:
                        del os.environ[v]
        if self._settings.get("request_timeout"):
            os.environ["GVC_REQUEST_TIMEOUT"] = str(self._settings["request_timeout"])
        if self._settings.get("max_game_workers"):
            os.environ["GVC_MAX_GAME_WORKERS"] = str(self._settings["max_game_workers"])
        if self._settings.get("download_manager"):
            os.environ["GVC_DOWNLOAD_MANAGER"] = self._settings["download_manager"]
        if self._settings.get("download_dir"):
            os.environ["GVC_DOWNLOAD_DIR"] = self._settings["download_dir"]

    def _apply_settings(self, new_settings: dict) -> None:
        self._settings.update(new_settings)
        save_settings(self._settings)
        ctk.set_appearance_mode(self._settings.get("appearance_mode", "System"))
        ctk.set_default_color_theme(self._settings.get("color_theme", "blue"))
        os.environ["GVC_REQUEST_TIMEOUT"] = str(self._settings.get("request_timeout", 10))
        os.environ["GVC_MAX_GAME_WORKERS"] = str(self._settings.get("max_game_workers", 3))
        for key, custom_env in [("http_proxy", "GVC_HTTP_PROXY"), ("https_proxy", "GVC_HTTPS_PROXY")]:
            val = self._settings.get(key, "")
            std_env = "HTTP_PROXY" if key == "http_proxy" else "HTTPS_PROXY"
            if val:
                os.environ[custom_env] = val
                os.environ[std_env] = val  # google-play-scraper 依赖标准变量
            else:
                for v in (custom_env, std_env):
                    if v in os.environ:
                        del os.environ[v]
        if self._settings.get("download_manager"):
            os.environ["GVC_DOWNLOAD_MANAGER"] = self._settings["download_manager"]
        if self._settings.get("download_dir"):
            os.environ["GVC_DOWNLOAD_DIR"] = self._settings["download_dir"]
        self._status_label.configure(
            text=" 设置已保存" + (" | 代理已配置" if new_settings.get("http_proxy") else "")
        )

    def _on_about(self) -> None:
        AboutDialog(self)

    def _on_recheck_selected(self) -> None:
        pkg = self._table.get_selected_package()
        if not pkg:
            messagebox.showwarning("提示", "请先在表格中右键选择一款游戏")
            return
        dialog = SingleCheckDialog(self)
        dialog._pkg_var.set(pkg)

    def _update_mode_indicator(self) -> None:
        self._mode_label.configure(text=f" {ctk.get_appearance_mode()} mode ")

    def _set_status_dot(self, color: str) -> None:
        """设置状态圆点颜色: muted / active / success / error."""
        colors = {
            "muted":   ("#64748B", "#94A3B8"),
            "active":  ("#3B82F6", "#60A5FA"),
            "success": ("#22C55E", "#4ADE80"),
            "error":   ("#EF4444", "#F87171"),
        }
        self._status_dot.configure(text_color=colors.get(color, colors["muted"]))

    def _on_appearance_changed(self) -> None:
        self._update_mode_indicator()
        self._build_menu()
        # 刷新 ttk Treeview 样式
        from gvc_gui.main_view import _apply_ttk_style
        _apply_ttk_style()

    def _on_close(self) -> None:
        if self._running:
            if not messagebox.askyesno("确认", "排查正在进行中，确定要退出吗？"):
                return
            self._stop_check()
        self.destroy()


def main() -> None:
    app = MainApplication()
    app.mainloop()


if __name__ == "__main__":
    main()
